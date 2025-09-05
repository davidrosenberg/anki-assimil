import mutagen
from mutagen.mp3 import MP3

fname = "/Users/drosen/Desktop/assimil course/L016-Hebrew ASSIMIL/S02.mp3"
audio = MP3(fname)
print(audio.info.length)
print(audio.info.bitrate)

from mutagen.id3 import ID3, TIT2
mutagen.File(fname).keys()

from mutagen.easyid3 import EasyID3
audio = MP3(fname, ID3=EasyID3)
print(audio.pprint()  )

import sys
sys.path.append('/usr/share/anki')
from anki import Collection
from anki.utils import intTime
import time

collection_path = "/home/username/Documents/Anki/User 1/collection.anki2"
col = Collection(collection_path)

tody = intTime() #anki function returns current integer time
nextTwenty = today + 20*60 #integer time in 20 minutes

query="select count(id) from cards where queue = 1 and due < %s" % nextTwenty
learnAheadCards = col.db.scalar(query)

print("You have %s learning cards due in Anki" % learnAheadCards)

col.close()



-----
from mutagen.easyid3 import EasyID3
fname = "/Users/drosen/Desktop/assimil course/L016-Hebrew ASSIMIL/S03.mp3"
audio = MP3(fname, ID3=EasyID3)
print(audio[['title']])

import shutil
import os
from googletrans import Translator
translator = Translator()
directory = "/Users/drosen/Desktop/assimil course/L018-Hebrew ASSIMIL"
mediafolder = "/Users/drosen/Desktop/anki-media-import"
rows = []
for filename in os.listdir(directory):
    fname = os.path.join(directory, filename)
    #print(fname)
    audio = MP3(fname, ID3=EasyID3)
    #print(audio['title'][0].split('-'))
    lesson = audio['album'][0].split(' - ')[-1]
    split_title = audio['title'][0].split('-')
    hebrew_text = split_title[-1] #.rstrip('٭')
    trans = translator.translate(hebrew_text)
    id = '.'.join([lesson, split_title[0]])
    tags = ' '.join(['assimil', lesson])
    #print(lesson, split_title[0], hebrew_text )
    newfname = '.'.join([id,'mp3'])
    newfullpath = os.path.join(mediafolder,newfname)
    shutil.copyfile(fname, newfullpath)
    row = (id, trans.origin, trans.text, ''.join(['[sound:',newfname,']']), tags )
    rows.append(row)


import csv

with open('/Users/drosen/Desktop/assimil.csv', 'w', newline='', encoding='utf-8') as csv_file:
    writer = csv.writer(csv_file, delimiter=';')
    writer.writerows(rows)

#  cp  ~/Desktop/anki-media-import/* '/Users/drosen/Library/Application Support/Anki2/User 1/collection.media/'
# media folder: /Users/drosen/Library/Application\ Support/Anki2/User\ 1/collection.media


import genanki

my_model = genanki.Model(
  1607392319,
  'Simple Model',
  fields=[
    {'name': 'Question'},
     {'name': 'Answer'},
  ],
  templates=[
    {
      'name': 'Card 1',
      'qfmt': '{{Question}}',
      'afmt': '{{FrontSide}}<hr id="answer">{{Answer}}',
    },
  ])

my_note = genanki.Note(
  model=my_model,
  fields=['Capital of Argentina', 'Buenos Aires'])
#####

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
